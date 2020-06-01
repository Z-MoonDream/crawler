import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook

# 读取工作表
# from openpyxl import load_workbook
# book = load_workbook("豆瓣电影TOP250.xlsx")
# 原地址
url = "https://movie.douban.com/top250?start="
# 保存数据
book = Workbook()
sheet = book.create_sheet("豆瓣电影TOP250", 0)
hbooklist = ["排名", "名称", "电影图片", "简单介绍", "评分", "评价人数", "简评"]
for i in range(0, 7):
    sheet.cell(1, i + 1).value = hbooklist[i]
# 标题头模拟浏览器
header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/50.0.2661.102 UBrowser/6.1.2107.204 Safari/537.36',
    # 'Referer': 'https://www.douban.com/',
}


# 爬取网页
def req(url, header):
    try:
        html = requests.get(url=url, headers=header)
    except EOFError as e:
        print(e)
    return (html)


# 解析数据,并保存
def analysis_save(html):
    soup = BeautifulSoup(html.text, "lxml")
    ranking = [soup.find_all("em")[i].text for i in range(25)]
    title = [soup.find_all("img")[i]["alt"] for i in range(25)]
    img = [soup.find_all("img")[i]["src"] for i in range(25)]
    introduce = []
    for index in soup.find_all("p", class_=""):
        introduce.append("".join(index.text.split()))  # 因为会出现xa0的空格符，用split分割后在转化成字符串去除
    grade = [soup.find_all("span", class_="rating_num")[i].text for i in range(25)]
    people = [soup.find_all(string=re.compile("人评价"))[i] for i in range(25)]
    # 简评中有空值，并且还是没有span标签的那种，只能做判断了，如果没有简评就插入""要不后面加入工作簿会报错(耦合性太高了)
    comment = []
    if len(soup.find_all("span", class_="inq")) == 25:
        comment = [soup.find_all("span", class_="inq")[i].text for i in range(25)]
    else:
        for d in soup.find_all("div", class_="item"):
            com = d.find("span", class_="inq")
            if com == None:
                comment.append("")
            else:
                comment.append(com.text)
    tup = (ranking, title, img, introduce, grade, people, comment)
    # 加入工作簿
    ws = book[book.sheetnames[0]]
    n = 0
    for row in range(int(ranking[0]) + 1, int(ranking[-1]) + 2):
        # n += 1  在这里设置的话n进来是1，最终n是26会爆要在加入数据之后加1
        for column in range(1, 8):
            ws.cell(row, column).value = tup[column - 1][n]
        n += 1


def main():
    for i in range(10):
        url = "https://movie.douban.com/top250?start=" + str(i * 25)
        html = req(url, header)
        analysis_save(html)
    # 保存
    book.save("豆瓣电影TOP250.xlsx")


if __name__ == '__main__':
    main()
