from openpyxl import Workbook
from openpyxl import load_workbook

# book = Workbook()

# sheet= book.create_sheet("newsheet",0)

#  打开已有工作簿

sheet = load_workbook("text.xlsx") #[工作表名字就可以打开了]
ws = sheet[sheet.sheetnames[0]]

# 修改

# 用cell修改
# for x in range(1,10,):
#     for j in range(1,11-x):
#         ws.cell(x,j).value=f"{10-x}*{j}={x*j}"
                # 行，列
# 9*1=1	9*2=2	9*3=3	9*4=4	9*5=5	9*6=6	9*7=7	9*8=8	9*9=9
# 8*1=2	8*2=4	8*3=6	8*4=8	8*5=10	8*6=12	8*7=14	8*8=16
# 7*1=3	7*2=6	7*3=9	7*4=12	7*5=15	7*6=18	7*7=21
# 6*1=4	6*2=8	6*3=12	6*4=16	6*5=20	6*6=24
# 5*1=5	5*2=10	5*3=15	5*4=20	5*5=25




# 3*1=7	3*2=14	3*3=21
# 2*1=8	2*2=16
# 1*1=9




# 修改的另一种方法
# for x in ws["A1:G100"]: # A1----G1  一直到 A100 到G100
#                         #  前面字母会变成后面的字母
#                          #   前面数字会变成后面数字
#     for call in x:
#         print(x,call)
#         call.value=None
sheet.save("text.xlsx")
# 保存的时候是保存的打开的那个 比如 wb=Worbook() 就是 wb.save()
#                                   wb= load_workbook() 就是 wb.save()









